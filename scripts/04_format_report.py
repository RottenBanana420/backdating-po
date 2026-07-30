"""
Formats data/output/po_reporting_periods.xlsx (produced by
03_split_by_reporting_period.py) for human review:

  - Each worksheet is sorted by reporting_month ascending (oldest to
    newest).
  - Every row is shaded with a subtle background color keyed to its
    reporting_month; consecutive reporting_month sections alternate
    through a palette so each section is visually distinct.
  - The tlc and receive_date columns (the two values a reviewer
    compares to spot backdated POs) are made to stand out from the
    rest of the row: a fixed fill (HIGHLIGHT_FILL, replacing the
    inherited reporting_month row color), bold navy text, a solid navy
    header, and a bold, saturated border (HIGHLIGHT_BORDER_COLOR)
    around every cell. Both columns are trimmed to date-only (the time
    portion isn't meaningful for the comparison) and displayed with a
    clean, centered date format.
  - Every column gets a bold, centered, grey header with a bottom
    border, and is auto-sized to fit its contents. The header row is
    frozen so it stays visible while scrolling through the sorted data.
"""
import re
import zipfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "output" / "po_reporting_periods.xlsx"
DST = SRC

TEXT_COLUMNS = ["po", "invoiceid"]
HIGHLIGHT_COLUMNS = ["tlc", "receive_date"]

# Subtle, mutually distinguishable row-shading colors, cycled per
# reporting_month section.
MONTH_FILL_COLORS = [
    "DDEBF7",  # light blue
    "E2EFDA",  # light green
    "FFF2CC",  # light yellow
    "FCE4D6",  # light orange
    "EAD1DC",  # light pink
    "E4DFEC",  # light purple
    "D9E1F2",  # light indigo
    "F2F2F2",  # light grey
]

ACCENT_COLOR = "1F4E78"
HIGHLIGHT_FONT = Font(bold=True, color=ACCENT_COLOR)
HIGHLIGHT_DATE_FORMAT = "mm/dd/yyyy"
HIGHLIGHT_ALIGNMENT = Alignment(horizontal="center")
HEADER_HIGHLIGHT_FILL = PatternFill("solid", fgColor=ACCENT_COLOR)
HEADER_HIGHLIGHT_FONT = Font(bold=True, color="FFFFFF")

# Bold, highly saturated border color chosen to read clearly against every
# color in MONTH_FILL_COLORS (all light pastels) as well as the navy header.
HIGHLIGHT_BORDER_COLOR = "C00000"
# Fill for tlc/receive_date data cells: a fixed, distinct tint (not one of
# MONTH_FILL_COLORS) in the same red family as the border, so these two
# columns read as one consistent block regardless of the row's month color.
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FCE4E4")
HIGHLIGHT_BORDER = Border(
    left=Side(style="thick", color=HIGHLIGHT_BORDER_COLOR),
    right=Side(style="thick", color=HIGHLIGHT_BORDER_COLOR),
    top=Side(style="thick", color=HIGHLIGHT_BORDER_COLOR),
    bottom=Side(style="thick", color=HIGHLIGHT_BORDER_COLOR),
)

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FONT = Font(bold=True, color="000000")
HEADER_BORDER = Border(bottom=Side(style="thin", color="808080"))
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 40


def parse_reporting_month(value):
    return datetime.strptime(value, "%B %Y")


def to_date_only(value):
    # tlc / receive_date come in as timestamp strings, e.g.
    # "5/26/2026 13:03:07:59" / "5/22/2026 00:00:00" - the time portion
    # isn't meaningful for the tlc-vs-receive_date comparison, so keep
    # only the date. Also accepts an already-converted date/datetime
    # (e.g. if this script is re-run against its own prior output).
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    date_part = str(value).split(" ")[0]
    return datetime.strptime(date_part, "%m/%d/%Y").date()


wb = load_workbook(SRC)

# sheet_name -> (text_column_letters, last_row), needed to re-patch the
# "Number Stored as Text" ignoredErrors block after save (openpyxl does
# not serialize it, matching the workaround in 03_split_by_reporting_period.py).
text_column_letters_by_sheet = {}

for ws in wb.worksheets:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= 1:
        continue
    header, data_rows = rows[0], rows[1:]

    reporting_month_idx = header.index("reporting_month")

    data_rows.sort(key=lambda r: parse_reporting_month(r[reporting_month_idx]))

    highlight_indices = {header.index(c) for c in HIGHLIGHT_COLUMNS}
    data_rows = [
        tuple(
            to_date_only(value) if col_num in highlight_indices else value
            for col_num, value in enumerate(row)
        )
        for row in data_rows
    ]

    for row_num, row in enumerate(data_rows, start=2):
        for col_num, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_num, value=value)

    # reporting_month / po / invoiceid must stay Text-formatted and
    # left-aligned (matches 02_build_report.py / 03_split_by_reporting_period.py).
    reporting_month_letter = get_column_letter(reporting_month_idx + 1)
    for cell in ws[reporting_month_letter][1:]:
        cell.number_format = "@"
        cell.alignment = Alignment(horizontal="left")

    text_column_letters = []
    for col_name in TEXT_COLUMNS:
        col_idx = header.index(col_name) + 1
        col_letter = get_column_letter(col_idx)
        text_column_letters.append(col_letter)
        for cell in ws[col_letter][1:]:
            cell.number_format = "@"
            cell.alignment = Alignment(horizontal="left")
    text_column_letters_by_sheet[ws.title] = (text_column_letters, ws.max_row)

    # Shade each full row by its reporting_month, alternating colors
    # each time the reporting_month changes.
    color_idx = -1
    prev_month = None
    for row_num, row in enumerate(data_rows, start=2):
        month = row[reporting_month_idx]
        if month != prev_month:
            color_idx = (color_idx + 1) % len(MONTH_FILL_COLORS)
            prev_month = month
        fill = PatternFill("solid", fgColor=MONTH_FILL_COLORS[color_idx])
        for col_num in range(1, len(header) + 1):
            ws.cell(row=row_num, column=col_num).fill = fill

    # Style every header cell (bold, centered, grey fill, bottom border)
    # and auto-size each column to fit its longest value.
    for col_idx, col_name in enumerate(header, start=1):
        col_letter = get_column_letter(col_idx)
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.fill = HEADER_FILL
        header_cell.font = HEADER_FONT
        header_cell.border = HEADER_BORDER
        header_cell.alignment = HEADER_ALIGNMENT

        longest = max(
            [len(str(col_name))]
            + [len(str(row[col_idx - 1])) for row in data_rows if row[col_idx - 1] is not None]
        )
        ws.column_dimensions[col_letter].width = max(
            MIN_COLUMN_WIDTH, min(longest + 2, MAX_COLUMN_WIDTH)
        )

    # Make tlc / receive_date stand out on top of the row shading. This
    # runs after the generic header styling above so it overrides it,
    # replacing the inherited reporting_month row color with a fixed
    # fill so these two columns stand out consistently in every row.
    for col_name in HIGHLIGHT_COLUMNS:
        col_idx = header.index(col_name) + 1
        col_letter = get_column_letter(col_idx)

        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.fill = HEADER_HIGHLIGHT_FILL
        header_cell.font = HEADER_HIGHLIGHT_FONT
        header_cell.border = HIGHLIGHT_BORDER

        for cell in ws[col_letter][1:]:
            cell.fill = HIGHLIGHT_FILL
            cell.font = HIGHLIGHT_FONT
            cell.border = HIGHLIGHT_BORDER
            cell.number_format = HIGHLIGHT_DATE_FORMAT
            cell.alignment = HIGHLIGHT_ALIGNMENT

    ws.freeze_panes = "A2"

wb.save(DST)

with zipfile.ZipFile(DST, "r") as zin:
    names = zin.namelist()
    contents = {name: zin.read(name) for name in names}

workbook_xml = contents["xl/workbook.xml"].decode("utf-8")
rels_xml = contents["xl/_rels/workbook.xml.rels"].decode("utf-8")

rid_to_target = {}
for rel_tag in re.findall(r"<Relationship\b[^>]*/>", rels_xml):
    rid = re.search(r'Id="(rId\d+)"', rel_tag).group(1)
    target = re.search(r'Target="([^"]+)"', rel_tag).group(1)
    rid_to_target[rid] = target

name_to_rid = {}
for sheet_tag in re.findall(r"<sheet\b[^>]*/>", workbook_xml):
    name = re.search(r'name="([^"]+)"', sheet_tag).group(1)
    rid = re.search(r'r:id="(rId\d+)"', sheet_tag).group(1)
    name_to_rid[name] = rid

for sheet_name, (text_column_letters, last_row) in text_column_letters_by_sheet.items():
    target = rid_to_target[name_to_rid[sheet_name]].lstrip("/")
    sheet_path = target if target.startswith("xl/") else f"xl/{target}"

    ignored_errors_xml = "<ignoredErrors>" + "".join(
        f'<ignoredError sqref="{letter}2:{letter}{last_row}" numberStoredAsText="1"/>'
        for letter in text_column_letters
    ) + "</ignoredErrors>"

    sheet_xml = contents[sheet_path].decode("utf-8")
    sheet_xml = sheet_xml.replace("</worksheet>", ignored_errors_xml + "</worksheet>")
    contents[sheet_path] = sheet_xml.encode("utf-8")

with zipfile.ZipFile(DST, "w", zipfile.ZIP_DEFLATED) as zout:
    for name in names:
        zout.writestr(name, contents[name])

print(f"Sorted and formatted: {DST}")
for ws in wb.worksheets:
    print(f"  {ws.title}: {ws.max_row - 1} rows")
