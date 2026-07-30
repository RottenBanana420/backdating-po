"""
Trims raw_data_fixed.csv down to a fixed set of columns and adds a
reporting_month column (derived from receive_date, formatted "Month Year")
at the beginning of the dataset. Output is written as .xlsx with
reporting_month explicitly set to Text format and left-aligned, so Excel
does not auto-detect it as a date (which would right-align it) and no
apostrophe or other stray character is needed in the data itself.
"""
import csv
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "processed" / "raw_data_fixed.csv"
DST = ROOT / "data" / "output" / "raw_data_fixed.xlsx"
ENCODING = "cp1252"

columns_to_keep = [
    "vend",
    "reqdate",
    "po",
    "dlvdate",
    "invoiceid",
    "tlc",
    "receive_date",
    "rcvtotalamt",
    "dept",
    "rcv_tlc",
    "rcv_firstname",
    "rcv_lastname",
]

with open(SRC, encoding=ENCODING, newline="") as f:
    reader = csv.DictReader(f)
    rows = list(reader)

output_columns = ["reporting_month"] + columns_to_keep

out_rows = []
for row in rows:
    receive_date = datetime.strptime(row["receive_date"], "%m/%d/%Y %H:%M:%S")
    reporting_month = receive_date.strftime("%B %Y")
    new_row = [reporting_month]
    for col in columns_to_keep:
        value = row[col]
        if col == "rcvtotalamt":
            value = float(value)
        new_row.append(value)
    out_rows.append(new_row)

wb = Workbook()
ws = wb.active
ws.title = "raw_data"

ws.append(output_columns)

for row in out_rows:
    ws.append(row)

# Force the reporting_month column to Text format and left alignment
# so Excel does not auto-detect "May 2026" style values as dates.
for cell in ws["A"][1:]:
    cell.number_format = "@"
    cell.alignment = Alignment(horizontal="left")

# po and invoiceid: keep as text (already stored as strings), but suppress
# Excel's "Number Stored as Text" green-flag warning for those columns.
text_columns = ["po", "invoiceid"]
last_row = ws.max_row
text_column_letters = []
for col_name in text_columns:
    col_idx = output_columns.index(col_name) + 1
    col_letter = ws.cell(row=1, column=col_idx).column_letter
    text_column_letters.append(col_letter)
    for cell in ws[col_letter][1:]:
        cell.number_format = "@"
        cell.alignment = Alignment(horizontal="left")

wb.save(DST)

# openpyxl does not serialize IgnoredError on write, so patch the sheet XML
# directly to add an <ignoredErrors> block suppressing the "Number Stored
# as Text" flag for the po and invoiceid ranges.
ignored_errors_xml = "<ignoredErrors>" + "".join(
    f'<ignoredError sqref="{letter}2:{letter}{last_row}" numberStoredAsText="1"/>'
    for letter in text_column_letters
) + "</ignoredErrors>"

sheet_path = "xl/worksheets/sheet1.xml"
with zipfile.ZipFile(DST, "r") as zin:
    names = zin.namelist()
    contents = {name: zin.read(name) for name in names}

sheet_xml = contents[sheet_path].decode("utf-8")
sheet_xml = sheet_xml.replace("</worksheet>", ignored_errors_xml + "</worksheet>")
contents[sheet_path] = sheet_xml.encode("utf-8")

with zipfile.ZipFile(DST, "w", zipfile.ZIP_DEFLATED) as zout:
    for name in names:
        zout.writestr(name, contents[name])

print("Dataset successfully modified!")
print(f"Total rows: {len(out_rows)}")
print("\nColumn order:")
for i, col in enumerate(output_columns, 1):
    print(f"  {i}. {col}")
print(f"\nWritten to: {DST}")
