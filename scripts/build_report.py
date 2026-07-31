"""
Single-pass PO reporting pipeline.

Reads data/raw/raw_data.csv and writes the one final report directly to
data/output/po_reporting_periods.xlsx. No intermediate files are written
to disk; everything below happens in memory, in stages:

  1. load_and_clean_rows   - repairs rows torn in two by stray CR/LF
                              characters embedded (unquoted) inside the
                              `vendoraccountid` field.
  2. build_trimmed_rows    - trims to a fixed set of columns and adds a
                              reporting_month column (derived from
                              receive_date, formatted "Month Year").
  3. split_by_reporting_period - compares tlc vs receive_date per row and
                              buckets rows into "POs Within Reporting
                              Month" / "POs Outside Reporting Month"
                              (dropping rows where the two dates match).
  4. build_workbook        - sorts each sheet by reporting_month, shades
                              rows by month, highlights the tlc/
                              receive_date columns, adds a real Excel
                              Table + reporting_month Slicer, and returns
                              the finished workbook.
"""
import csv
import re
import zipfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

from _excel_slicers import SheetSlicerTarget, add_reporting_month_slicers

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "raw" / "raw_data.csv"
DST = ROOT / "data" / "output" / "po_reporting_periods.xlsx"
ENCODING = "cp1252"

COLUMNS_TO_KEEP = [
    "vend",
    "reqdate",
    "po",
    "dlvdate",
    "invoiceid",
    "tlc",
    "receive_date",
    "rcvtotalamt",
    "dept",
    "rcv_tlc",
    "rcv_firstname",
    "rcv_lastname",
]
TEXT_COLUMNS = ["po", "invoiceid"]
HIGHLIGHT_COLUMNS = ["tlc", "receive_date"]

WITHIN_SHEET = "POs Within Reporting Month"
OUTSIDE_SHEET = "POs Outside Reporting Month"

# Subtle, mutually distinguishable row-shading colors, cycled per
# reporting_month section.
MONTH_FILL_COLORS = [
    "DDEBF7",  # light blue
    "E2EFDA",  # light green
    "FFF2CC",  # light yellow
    "FCE4D6",  # light orange
    "EAD1DC",  # light pink
    "E4DFEC",  # light purple
    "D9E1F2",  # light indigo
    "F2F2F2",  # light grey
]

ACCENT_COLOR = "1F4E78"
HIGHLIGHT_FONT = Font(bold=True, color=ACCENT_COLOR)
HIGHLIGHT_DATE_FORMAT = "mm/dd/yyyy"
HIGHLIGHT_ALIGNMENT = Alignment(horizontal="center")
HEADER_HIGHLIGHT_FILL = PatternFill("solid", fgColor=ACCENT_COLOR)
HEADER_HIGHLIGHT_FONT = Font(bold=True, color="FFFFFF")

# Bold, highly saturated border color chosen to read clearly against every
# color in MONTH_FILL_COLORS (all light pastels) as well as the navy header.
HIGHLIGHT_BORDER_COLOR = "C00000"
# Fill for tlc/receive_date data cells: a fixed, distinct tint (not one of
# MONTH_FILL_COLORS) in the same red family as the border, so these two
# columns read as one consistent block regardless of the row's month color.
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FCE4E4")
HIGHLIGHT_BORDER = Border(
    left=Side(style="thick", color=HIGHLIGHT_BORDER_COLOR),
    right=Side(style="thick", color=HIGHLIGHT_BORDER_COLOR),
    top=Side(style="thick", color=HIGHLIGHT_BORDER_COLOR),
    bottom=Side(style="thick", color=HIGHLIGHT_BORDER_COLOR),
)

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FONT = Font(bold=True, color="000000")
HEADER_BORDER = Border(bottom=Side(style="thin", color="808080"))
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 40


def load_and_clean_rows(path):
    with open(path, encoding=ENCODING, newline="") as f:
        raw = f.read()

    # Split into logical lines on the file's real record delimiter only.
    lines = [ln for ln in raw.split("\r\n") if ln.strip() != ""]

    header = None
    reader_rows = []
    bad_rows = []

    for i, line in enumerate(lines, start=1):
        # Strip any stray bare \r or \n that survived within the line
        # (these are the characters that caused rows to split early).
        cleaned = line.replace("\r", "").replace("\n", "")
        fields = next(csv.reader([cleaned]))
        if header is None:
            header = fields
            continue
        if len(fields) != len(header):
            bad_rows.append((i, len(fields), fields))
            continue
        reader_rows.append(fields)

    print(f"Header columns: {len(header)}")
    print(f"Clean data rows: {len(reader_rows)}")
    print(f"Rows still misaligned after fix: {len(bad_rows)}")
    for i, n, fields in bad_rows[:20]:
        print(f"  line {i}: {n} fields -> {fields}")

    return header, reader_rows


def build_trimmed_rows(header, rows):
    output_columns = ["reporting_month"] + COLUMNS_TO_KEEP
    col_idx = {name: header.index(name) for name in COLUMNS_TO_KEEP}

    out_rows = []
    for row in rows:
        receive_date = datetime.strptime(row[col_idx["receive_date"]], "%m/%d/%Y %H:%M:%S")
        reporting_month = receive_date.strftime("%B %Y")
        new_row = [reporting_month]
        for col in COLUMNS_TO_KEEP:
            value = row[col_idx[col]]
            if col == "rcvtotalamt":
                value = float(value)
            new_row.append(value)
        out_rows.append(new_row)

    return output_columns, out_rows


def parse_date(value):
    date_part = value.split(" ")[0]
    return datetime.strptime(date_part, "%m/%d/%Y").date()


def split_by_reporting_period(header, rows):
    tlc_idx = header.index("tlc")
    receive_idx = header.index("receive_date")

    within_rows, outside_rows = [], []
    deleted_count = 0

    for row in rows:
        tlc_date = parse_date(row[tlc_idx])
        receive_date = parse_date(row[receive_idx])

        if tlc_date == receive_date:
            deleted_count += 1
            continue

        if (tlc_date.year, tlc_date.month) == (receive_date.year, receive_date.month):
            within_rows.append(row)
        else:
            outside_rows.append(row)

    print(f"Deleted (tlc == receive_date): {deleted_count}")
    print(f"{WITHIN_SHEET}: {len(within_rows)} rows")
    print(f"{OUTSIDE_SHEET}: {len(outside_rows)} rows")

    return {WITHIN_SHEET: within_rows, OUTSIDE_SHEET: outside_rows}


def parse_reporting_month(value):
    return datetime.strptime(value, "%B %Y")


def to_date_only(value):
    # tlc / receive_date come in as timestamp strings, e.g.
    # "5/26/2026 13:03:07:59" / "5/22/2026 00:00:00" - the time portion
    # isn't meaningful for the tlc-vs-receive_date comparison, so keep
    # only the date.
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    date_part = str(value).split(" ")[0]
    return datetime.strptime(date_part, "%m/%d/%Y").date()


def build_workbook(header, sheets):
    wb = Workbook()
    wb.remove(wb.active)

    reporting_month_idx = header.index("reporting_month")
    highlight_indices = {header.index(c) for c in HIGHLIGHT_COLUMNS}

    # sheet_name -> (text_column_letters, last_row), needed to re-patch the
    # "Number Stored as Text" ignoredErrors block after save (openpyxl does
    # not serialize it).
    text_column_letters_by_sheet = {}
    # sheet_name -> SheetSlicerTarget-ish info, needed to patch in a real
    # reporting_month Slicer after save (openpyxl can't create one itself -
    # see _excel_slicers.py).
    slicer_info_by_sheet = {}

    for table_id, (sheet_name, rows) in enumerate(sheets.items(), start=1):
        ws = wb.create_sheet(sheet_name)
        ws.append(list(header))

        data_rows = sorted(rows, key=lambda r: parse_reporting_month(r[reporting_month_idx]))
        data_rows = [
            tuple(
                to_date_only(value) if col_num in highlight_indices else value
                for col_num, value in enumerate(row)
            )
            for row in data_rows
        ]
        for row in data_rows:
            ws.append(list(row))

        # reporting_month / po / invoiceid must stay Text-formatted and
        # left-aligned so Excel doesn't auto-detect them and re-flag them.
        reporting_month_letter = get_column_letter(reporting_month_idx + 1)
        for cell in ws[reporting_month_letter][1:]:
            cell.number_format = "@"
            cell.alignment = Alignment(horizontal="left")

        text_column_letters = []
        for col_name in TEXT_COLUMNS:
            col_idx = header.index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            text_column_letters.append(col_letter)
            for cell in ws[col_letter][1:]:
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="left")
        text_column_letters_by_sheet[ws.title] = (text_column_letters, ws.max_row)

        # Shade each full row by its reporting_month, alternating colors
        # each time the reporting_month changes.
        color_idx = -1
        prev_month = None
        for row_num, row in enumerate(data_rows, start=2):
            month = row[reporting_month_idx]
            if month != prev_month:
                color_idx = (color_idx + 1) % len(MONTH_FILL_COLORS)
                prev_month = month
            fill = PatternFill("solid", fgColor=MONTH_FILL_COLORS[color_idx])
            for col_num in range(1, len(header) + 1):
                ws.cell(row=row_num, column=col_num).fill = fill

        # Style every header cell (bold, centered, grey fill, bottom border)
        # and auto-size each column to fit its longest value.
        for col_idx, col_name in enumerate(header, start=1):
            col_letter = get_column_letter(col_idx)
            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.fill = HEADER_FILL
            header_cell.font = HEADER_FONT
            header_cell.border = HEADER_BORDER
            header_cell.alignment = HEADER_ALIGNMENT

            longest = max(
                [len(str(col_name))]
                + [len(str(row[col_idx - 1])) for row in data_rows if row[col_idx - 1] is not None]
            )
            ws.column_dimensions[col_letter].width = max(
                MIN_COLUMN_WIDTH, min(longest + 2, MAX_COLUMN_WIDTH)
            )

        # Make tlc / receive_date stand out on top of the row shading. This
        # runs after the generic header styling above so it overrides it,
        # replacing the inherited reporting_month row color with a fixed
        # fill so these two columns stand out consistently in every row.
        for col_name in HIGHLIGHT_COLUMNS:
            col_idx = header.index(col_name) + 1
            col_letter = get_column_letter(col_idx)

            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.fill = HEADER_HIGHLIGHT_FILL
            header_cell.font = HEADER_HIGHLIGHT_FONT
            header_cell.border = HIGHLIGHT_BORDER

            for cell in ws[col_letter][1:]:
                cell.fill = HIGHLIGHT_FILL
                cell.font = HIGHLIGHT_FONT
                cell.border = HIGHLIGHT_BORDER
                cell.number_format = HIGHLIGHT_DATE_FORMAT
                cell.alignment = HIGHLIGHT_ALIGNMENT

        ws.freeze_panes = "A2"

        # Real Excel Table backing the reporting_month Slicer added below.
        # No built-in style/banding, so it doesn't fight with the month
        # shading and header styling already applied above.
        table_ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"
        table = Table(
            id=table_id,
            displayName=f"Table{table_id}",
            ref=table_ref,
            tableColumns=[TableColumn(id=i, name=str(name)) for i, name in enumerate(header, start=1)],
        )
        table.tableStyleInfo = TableStyleInfo(
            showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False
        )
        ws.add_table(table)
        slicer_info_by_sheet[ws.title] = (table_id, reporting_month_idx + 1, len(header))

    return wb, text_column_letters_by_sheet, slicer_info_by_sheet


def save_workbook(wb, text_column_letters_by_sheet, slicer_info_by_sheet, dst):
    wb.save(dst)

    with zipfile.ZipFile(dst, "r") as zin:
        names = zin.namelist()
        contents = {name: zin.read(name) for name in names}

    workbook_xml = contents["xl/workbook.xml"].decode("utf-8")
    rels_xml = contents["xl/_rels/workbook.xml.rels"].decode("utf-8")

    rid_to_target = {}
    for rel_tag in re.findall(r"<Relationship\b[^>]*/>", rels_xml):
        rid = re.search(r'Id="(rId\d+)"', rel_tag).group(1)
        target = re.search(r'Target="([^"]+)"', rel_tag).group(1)
        rid_to_target[rid] = target

    name_to_rid = {}
    for sheet_tag in re.findall(r"<sheet\b[^>]*/>", workbook_xml):
        name = re.search(r'name="([^"]+)"', sheet_tag).group(1)
        rid = re.search(r'r:id="(rId\d+)"', sheet_tag).group(1)
        name_to_rid[name] = rid

    sheet_path_by_name = {}
    for sheet_name in name_to_rid:
        target = rid_to_target[name_to_rid[sheet_name]].lstrip("/")
        sheet_path_by_name[sheet_name] = target if target.startswith("xl/") else f"xl/{target}"

    for sheet_name, (text_column_letters, last_row) in text_column_letters_by_sheet.items():
        sheet_path = sheet_path_by_name[sheet_name]

        ignored_errors_xml = "<ignoredErrors>" + "".join(
            f'<ignoredError sqref="{letter}2:{letter}{last_row}" numberStoredAsText="1"/>'
            for letter in text_column_letters
        ) + "</ignoredErrors>"

        # <ignoredErrors> must precede <drawing>/<tableParts> per the OOXML
        # worksheet element order, so anchor on <tableParts> (added by
        # ws.add_table() above) rather than blindly prepending </worksheet>.
        sheet_xml = contents[sheet_path].decode("utf-8")
        if "<tableParts" in sheet_xml:
            sheet_xml = sheet_xml.replace("<tableParts", ignored_errors_xml + "<tableParts", 1)
        else:
            sheet_xml = sheet_xml.replace("</worksheet>", ignored_errors_xml + "</worksheet>")
        contents[sheet_path] = sheet_xml.encode("utf-8")

    slicer_targets = [
        SheetSlicerTarget(
            sheet_path=sheet_path_by_name[sheet_name],
            table_id=table_id,
            slicer_index=table_id,
            reporting_month_column=reporting_month_column,
            num_columns=num_columns,
        )
        for sheet_name, (table_id, reporting_month_column, num_columns) in slicer_info_by_sheet.items()
    ]
    contents = add_reporting_month_slicers(contents, slicer_targets)

    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in contents:
            zout.writestr(name, contents[name])


def main():
    raw_header, raw_rows = load_and_clean_rows(SRC)
    header, rows = build_trimmed_rows(raw_header, raw_rows)
    sheets = split_by_reporting_period(header, rows)
    wb, text_column_letters_by_sheet, slicer_info_by_sheet = build_workbook(header, sheets)
    save_workbook(wb, text_column_letters_by_sheet, slicer_info_by_sheet, DST)

    print(f"\nWritten to: {DST}")
    for sheet_name, sheet_rows in sheets.items():
        print(f"  {sheet_name}: {len(sheet_rows)} rows")


if __name__ == "__main__":
    main()
