"""
Compares the tlc and receive_date columns for each row of
data/output/raw_data_fixed.xlsx (produced by 02_build_report.py) and
splits the rows as follows:

  - tlc date == receive_date date       -> row dropped
  - tlc date and receive_date date fall
    in the same month and year         -> "POs Within Reporting Month"
  - tlc date and receive_date date fall
    in different months and/or years   -> "POs Outside Reporting Month"

Only the date portion of tlc/receive_date is compared (both columns are
timestamps like "5/6/2026 17:12:49:983" / "5/6/2026 00:00:00"); the time
portion is ignored.

Output is written as a new workbook containing exactly these two
worksheets.
"""
import re
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "output" / "raw_data_fixed.xlsx"
DST = ROOT / "data" / "output" / "po_reporting_periods.xlsx"

WITHIN_SHEET = "POs Within Reporting Month"
OUTSIDE_SHEET = "POs Outside Reporting Month"

# po and invoiceid: keep as text (already stored as strings), but suppress
# Excel's "Number Stored as Text" green-flag warning for those columns
# (matches 02_build_report.py).
TEXT_COLUMNS = ["po", "invoiceid"]


def parse_date(value):
    date_part = value.split(" ")[0]
    return datetime.strptime(date_part, "%m/%d/%Y").date()


src_wb = load_workbook(SRC)
src_ws = src_wb["raw_data"]

rows = list(src_ws.iter_rows(values_only=True))
header, data_rows = rows[0], rows[1:]

tlc_idx = header.index("tlc")
receive_idx = header.index("receive_date")
reporting_month_idx = header.index("reporting_month")

within_rows = []
outside_rows = []
deleted_count = 0

for row in data_rows:
    tlc_date = parse_date(row[tlc_idx])
    receive_date = parse_date(row[receive_idx])

    if tlc_date == receive_date:
        deleted_count += 1
        continue

    if (tlc_date.year, tlc_date.month) == (receive_date.year, receive_date.month):
        within_rows.append(row)
    else:
        outside_rows.append(row)

wb = Workbook()
wb.remove(wb.active)

# sheet_name -> list of column letters that need the ignoredErrors patch
text_column_letters_by_sheet = {}

for sheet_name, sheet_rows in (
    (WITHIN_SHEET, within_rows),
    (OUTSIDE_SHEET, outside_rows),
):
    ws = wb.create_sheet(sheet_name)
    ws.append(list(header))
    for row in sheet_rows:
        ws.append(list(row))

    # Preserve reporting_month as Text, left-aligned (matches 02_build_report.py).
    for cell in ws[ws.cell(row=1, column=reporting_month_idx + 1).column_letter][1:]:
        cell.number_format = "@"
        cell.alignment = Alignment(horizontal="left")

    text_column_letters = []
    for col_name in TEXT_COLUMNS:
        col_idx = header.index(col_name) + 1
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        text_column_letters.append(col_letter)
        for cell in ws[col_letter][1:]:
            cell.number_format = "@"
            cell.alignment = Alignment(horizontal="left")
    text_column_letters_by_sheet[sheet_name] = (text_column_letters, ws.max_row)

wb.save(DST)

# openpyxl does not serialize IgnoredError on write, so patch each sheet's
# XML directly to add an <ignoredErrors> block suppressing the "Number
# Stored as Text" flag for the po and invoiceid ranges (matches
# 02_build_report.py). Map sheet name -> xl/worksheets/sheetN.xml via the
# workbook's own rels rather than assuming file order.
with zipfile.ZipFile(DST, "r") as zin:
    names = zin.namelist()
    contents = {name: zin.read(name) for name in names}

workbook_xml = contents["xl/workbook.xml"].decode("utf-8")
rels_xml = contents["xl/_rels/workbook.xml.rels"].decode("utf-8")

rid_to_target = {}
for rel_tag in re.findall(r"<Relationship\b[^>]*/>", rels_xml):
    rid = re.search(r'Id="(rId\d+)"', rel_tag).group(1)
    target = re.search(r'Target="([^"]+)"', rel_tag).group(1)
    rid_to_target[rid] = target

name_to_rid = {}
for sheet_tag in re.findall(r"<sheet\b[^>]*/>", workbook_xml):
    name = re.search(r'name="([^"]+)"', sheet_tag).group(1)
    rid = re.search(r'r:id="(rId\d+)"', sheet_tag).group(1)
    name_to_rid[name] = rid

for sheet_name, (text_column_letters, last_row) in text_column_letters_by_sheet.items():
    target = rid_to_target[name_to_rid[sheet_name]].lstrip("/")
    sheet_path = target if target.startswith("xl/") else f"xl/{target}"

    ignored_errors_xml = "<ignoredErrors>" + "".join(
        f'<ignoredError sqref="{letter}2:{letter}{last_row}" numberStoredAsText="1"/>'
        for letter in text_column_letters
    ) + "</ignoredErrors>"

    sheet_xml = contents[sheet_path].decode("utf-8")
    sheet_xml = sheet_xml.replace("</worksheet>", ignored_errors_xml + "</worksheet>")
    contents[sheet_path] = sheet_xml.encode("utf-8")

with zipfile.ZipFile(DST, "w", zipfile.ZIP_DEFLATED) as zout:
    for name in names:
        zout.writestr(name, contents[name])

print(f"Deleted (tlc == receive_date): {deleted_count}")
print(f"{WITHIN_SHEET}: {len(within_rows)} rows")
print(f"{OUTSIDE_SHEET}: {len(outside_rows)} rows")
print(f"\nWritten to: {DST}")
