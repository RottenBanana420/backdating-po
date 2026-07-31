from openpyxl import load_workbook

from src.config import COLUMNS_TO_KEEP, HIGHLIGHT_COLUMNS, OUTSIDE_SHEET, WITHIN_SHEET
from src.excel.styles import (
    HIGHLIGHT_BORDER,
    HIGHLIGHT_FILL,
    HIGHLIGHT_FONT,
    MAX_COLUMN_WIDTH,
    MIN_COLUMN_WIDTH,
    MONTH_FILL_COLORS,
)
from src.excel.workbook_builder import build_workbook

HEADER = ["reporting_month", *COLUMNS_TO_KEEP]


def _row(reporting_month="January 2026", tlc="1/15/2026", receive_date="1/20/2026 00:00:00",
         vend="Acme Co", po="PO123", dept="IT"):
    return [reporting_month, vend, "1/1/2026", po, "1/10/2026", "INV1", tlc,
            receive_date, 100.00, dept, "TLC1", "Jane", "Doe"]


def _build_and_reload(tmp_path, sheets):
    """build_workbook returns a write-only Workbook, which streams cells
    straight to disk and can't be read back in memory - so tests save it and
    re-open it with load_workbook() to inspect the actual written content,
    the same round-trip save_workbook does in production."""
    wb, text_cols, slicer_info = build_workbook(HEADER, sheets)
    path = tmp_path / "out.xlsx"
    wb.save(path)
    return load_workbook(path), text_cols, slicer_info


def test_build_workbook_creates_one_sheet_per_bucket_with_correct_names(tmp_path):
    sheets = {WITHIN_SHEET: [_row()], OUTSIDE_SHEET: [_row()]}

    wb, _, _ = _build_and_reload(tmp_path, sheets)

    assert wb.sheetnames == [WITHIN_SHEET, OUTSIDE_SHEET]


def test_build_workbook_shades_rows_by_reporting_month(tmp_path):
    rows = [
        _row(reporting_month="January 2026", receive_date="1/20/2026 00:00:00"),
        _row(reporting_month="January 2026", receive_date="1/21/2026 00:00:00"),
        _row(reporting_month="February 2026", receive_date="2/5/2026 00:00:00"),
    ]

    wb, _, _ = _build_and_reload(tmp_path, {WITHIN_SHEET: rows})
    ws = wb[WITHIN_SHEET]

    # Rows are sorted by reporting_month before writing, so rows 2-3 are
    # both January (same shading color) and row 4 is February (next color
    # in the MONTH_FILL_COLORS cycle).
    jan_fill_a = ws.cell(row=2, column=1).fill.fgColor.rgb
    jan_fill_b = ws.cell(row=3, column=1).fill.fgColor.rgb
    feb_fill = ws.cell(row=4, column=1).fill.fgColor.rgb

    assert jan_fill_a == jan_fill_b == "00" + MONTH_FILL_COLORS[0]
    assert feb_fill == "00" + MONTH_FILL_COLORS[1]


def test_build_workbook_highlights_tlc_and_receive_date_columns(tmp_path):
    wb, _, _ = _build_and_reload(tmp_path, {WITHIN_SHEET: [_row()]})
    ws = wb[WITHIN_SHEET]

    for col_name in HIGHLIGHT_COLUMNS:
        col_idx = HEADER.index(col_name) + 1
        cell = ws.cell(row=2, column=col_idx)
        assert cell.fill.fgColor.rgb == HIGHLIGHT_FILL.fgColor.rgb
        assert cell.font.bold == HIGHLIGHT_FONT.bold
        assert cell.font.color.rgb == HIGHLIGHT_FONT.color.rgb
        assert cell.border.left.style == HIGHLIGHT_BORDER.left.style
        assert cell.border.left.color.rgb == HIGHLIGHT_BORDER.left.color.rgb
        assert cell.number_format == "mm/dd/yyyy"


def test_build_workbook_attaches_table_with_correct_headers(tmp_path):
    wb, _, slicer_info = _build_and_reload(tmp_path, {WITHIN_SHEET: [_row()]})
    ws = wb[WITHIN_SHEET]

    assert "Table1" in ws.tables
    table = ws.tables["Table1"]
    assert [col.name for col in table.tableColumns] == HEADER
    assert slicer_info[WITHIN_SHEET] == (1, HEADER.index("reporting_month") + 1, len(HEADER))


def test_build_workbook_defuses_formula_injection_in_free_text_columns(tmp_path):
    rows = [
        _row(vend="=HYPERLINK(\"http://evil.example\",\"click\")", po="+PO123", dept="-IT"),
        _row(vend="@SUM(1,1)"),
    ]

    wb, _, _ = _build_and_reload(tmp_path, {WITHIN_SHEET: rows})
    ws = wb[WITHIN_SHEET]

    vend_col = HEADER.index("vend") + 1
    po_col = HEADER.index("po") + 1
    dept_col = HEADER.index("dept") + 1

    row1, row2 = ws[2], ws[3]

    # Value is left exactly as sourced - no apostrophe added - but marked
    # non-formula (data_type 's') and quote-prefixed so Excel renders and
    # re-edits it as plain text instead of evaluating it.
    assert row1[vend_col - 1].value == "=HYPERLINK(\"http://evil.example\",\"click\")"
    assert row1[vend_col - 1].data_type == "s"
    assert row1[vend_col - 1].quotePrefix is True
    assert row1[po_col - 1].value == "+PO123"
    assert row1[po_col - 1].quotePrefix is True
    assert row1[dept_col - 1].value == "-IT"
    assert row1[dept_col - 1].quotePrefix is True
    assert row2[vend_col - 1].value == "@SUM(1,1)"
    assert row2[vend_col - 1].quotePrefix is True


def test_build_workbook_leaves_ordinary_text_untouched(tmp_path):
    wb, _, _ = _build_and_reload(tmp_path, {WITHIN_SHEET: [_row(vend="Acme Co", po="PO123")]})
    ws = wb[WITHIN_SHEET]

    assert ws.cell(row=2, column=HEADER.index("vend") + 1).value == "Acme Co"
    assert ws.cell(row=2, column=HEADER.index("po") + 1).value == "PO123"


def test_build_workbook_sizes_columns_to_longest_value(tmp_path):
    rows = [
        _row(vend="Acme Co"),
        _row(vend="A Very Long Vendor Name That Exceeds The Max Width By A Lot"),
    ]

    wb, _, _ = _build_and_reload(tmp_path, {WITHIN_SHEET: rows})
    ws = wb[WITHIN_SHEET]

    vend_letter = ws.cell(row=1, column=HEADER.index("vend") + 1).column_letter
    dept_letter = ws.cell(row=1, column=HEADER.index("dept") + 1).column_letter

    # Longest "vend" value (60 chars) + 2 padding clamps to MAX_COLUMN_WIDTH.
    assert ws.column_dimensions[vend_letter].width == MAX_COLUMN_WIDTH
    # "dept" values ("IT") are shorter than the header itself, so width
    # falls back to MIN_COLUMN_WIDTH rather than shrinking further.
    assert ws.column_dimensions[dept_letter].width == MIN_COLUMN_WIDTH


def test_build_workbook_handles_empty_bucket_without_error(tmp_path):
    wb, text_cols, slicer_info = _build_and_reload(
        tmp_path, {WITHIN_SHEET: [], OUTSIDE_SHEET: [_row()]}
    )
    ws = wb[WITHIN_SHEET]
    ws_outside = wb[OUTSIDE_SHEET]

    # Verify empty bucket is handled correctly
    assert ws.max_row == 1  # header row only, no data rows
    assert ws.cell(row=1, column=1).value == "reporting_month"
    assert "Table1" in ws.tables
    assert WITHIN_SHEET in text_cols
    assert WITHIN_SHEET in slicer_info

    # Verify sibling non-empty bucket still builds correctly
    assert ws_outside.max_row == 2  # header + 1 data row
    assert "Table2" in ws_outside.tables
    assert OUTSIDE_SHEET in text_cols
    assert OUTSIDE_SHEET in slicer_info


def test_build_workbook_reports_progress_at_configured_interval(tmp_path):
    rows = [_row() for _ in range(3)]
    calls = []

    wb, _, _ = build_workbook(
        HEADER, {WITHIN_SHEET: rows}, on_progress=lambda done, total: calls.append((done, total)),
        progress_interval_rows=2,
    )
    # A write-only workbook's worksheets stream to an internal generator that
    # must be drained via save() - skipping this (as the two on_progress
    # tests originally did) leaves it to be torn down by the garbage
    # collector instead, at an arbitrary later point in the suite, which
    # openpyxl surfaces as a spurious "I/O operation on closed file" warning.
    wb.save(tmp_path / "out.xlsx")

    # Initial call before any rows, one at the row-2 interval boundary, and
    # a final flush at sheet end even though 3 isn't a multiple of 2 - so the
    # last call always lands exactly on the true total rather than stalling
    # one row short of it.
    assert calls == [(0, 3), (2, 3), (3, 3)]


def test_build_workbook_progress_totals_span_all_sheets(tmp_path):
    sheets = {WITHIN_SHEET: [_row(), _row()], OUTSIDE_SHEET: [_row()]}
    calls = []

    wb, _, _ = build_workbook(
        HEADER, sheets, on_progress=lambda done, total: calls.append((done, total)),
        progress_interval_rows=1,
    )
    wb.save(tmp_path / "out.xlsx")

    # rows_total (3) is fixed across the whole build, not reset per sheet,
    # and rows_done only ever increases, finishing at exactly rows_total.
    totals = {total for _, total in calls}
    assert totals == {3}
    done_values = [done for done, _ in calls]
    assert done_values == sorted(done_values)
    assert done_values[-1] == 3


def test_build_workbook_without_on_progress_behaves_unchanged(tmp_path):
    # on_progress omitted, as existing callers (CLI, tests) do - must not
    # raise just because no callback was given.
    wb, _, _ = _build_and_reload(tmp_path, {WITHIN_SHEET: [_row()]})
    assert wb[WITHIN_SHEET].max_row == 2
