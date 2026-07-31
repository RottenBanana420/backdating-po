import zipfile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.config import COLUMNS_TO_KEEP, TEXT_COLUMNS, WITHIN_SHEET
from src.excel.workbook_builder import build_workbook
from src.excel.workbook_writer import save_workbook

HEADER = ["reporting_month", *COLUMNS_TO_KEEP]


def _row(reporting_month="January 2026", tlc="1/15/2026", receive_date="1/20/2026 00:00:00",
         vend="Acme Co", po="PO123", dept="IT"):
    return [reporting_month, vend, "1/1/2026", po, "1/10/2026", "INV1", tlc,
            receive_date, 100.00, dept, "TLC1", "Jane", "Doe"]


def _build_and_save(dst, sheets):
    wb, text_cols, slicer_info = build_workbook(HEADER, sheets)
    save_workbook(wb, text_cols, slicer_info, dst)


def test_save_workbook_writes_a_valid_xlsx(tmp_path):
    dst = tmp_path / "out.xlsx"

    _build_and_save(dst, {WITHIN_SHEET: [_row()]})

    assert dst.is_file()
    with zipfile.ZipFile(dst) as zf:
        assert zf.testzip() is None  # None means every part's CRC checks out

    wb = load_workbook(dst)
    assert wb.sheetnames == [WITHIN_SHEET]


def test_save_workbook_adds_ignored_errors_for_text_columns(tmp_path):
    dst = tmp_path / "out.xlsx"

    _build_and_save(dst, {WITHIN_SHEET: [_row(), _row()]})

    with zipfile.ZipFile(dst) as zf:
        sheet_path = next(n for n in zf.namelist() if n.startswith("xl/worksheets/sheet"))
        sheet_xml = zf.read(sheet_path).decode("utf-8")

    assert "<ignoredErrors>" in sheet_xml
    assert sheet_xml.count("<ignoredError ") == len(TEXT_COLUMNS)
    for col_name in TEXT_COLUMNS:
        letter = get_column_letter(HEADER.index(col_name) + 1)
        assert f'sqref="{letter}2:{letter}3" numberStoredAsText="1"' in sheet_xml


def test_save_workbook_writes_to_the_exact_destination_path(tmp_path):
    dst = tmp_path / "nested_name.xlsx"

    _build_and_save(dst, {WITHIN_SHEET: [_row()]})

    assert dst.is_file()
    assert list(tmp_path.glob("*.xlsx")) == [dst]


def test_save_workbook_injects_reporting_month_slicer_parts(tmp_path):
    dst = tmp_path / "out.xlsx"

    _build_and_save(dst, {WITHIN_SHEET: [_row()]})

    with zipfile.ZipFile(dst) as zf:
        names = zf.namelist()

    assert "xl/slicerCaches/slicerCache1.xml" in names
    assert "xl/slicers/slicer1.xml" in names
    assert "xl/drawings/drawing1.xml" in names


def test_save_workbook_stores_unchanged_parts_without_recompressing(tmp_path):
    dst = tmp_path / "out.xlsx"

    _build_and_save(dst, {WITHIN_SHEET: [_row()]})

    with zipfile.ZipFile(dst) as zf:
        info_by_name = {info.filename: info for info in zf.infolist()}

    # openpyxl writes these parts and save_workbook never touches them - no
    # need to spend CPU re-deflating bytes that are already on disk verbatim.
    for unchanged in ("xl/styles.xml", "xl/theme/theme1.xml", "docProps/core.xml"):
        assert info_by_name[unchanged].compress_type == zipfile.ZIP_STORED

    # Sheet XML always gets an <ignoredErrors> and/or slicer patch, so it
    # must stay compressed rather than bloating the file.
    sheet_path = next(n for n in info_by_name if n.startswith("xl/worksheets/sheet"))
    assert info_by_name[sheet_path].compress_type == zipfile.ZIP_DEFLATED
