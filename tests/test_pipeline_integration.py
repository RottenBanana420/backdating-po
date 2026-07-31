from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.pipeline import run

FIXTURE = Path(__file__).parent.parent / "data" / "sample" / "sample_raw_data.csv"
ALL_DROPPED_FIXTURE = Path(__file__).parent / "fixtures" / "all_dropped_raw_data.csv"


def test_pipeline_end_to_end(tmp_path):
    dst = tmp_path / "out.xlsx"
    counts = run(src=FIXTURE, dst=dst)

    assert dst.exists()
    assert sum(counts.values()) > 0

    wb = load_workbook(dst)
    assert set(wb.sheetnames) <= {"POs Within Reporting Month", "POs Outside Reporting Month"}
    for name in wb.sheetnames:
        ws = wb[name]
        assert ws.cell(row=2, column=1).value == "reporting_month"


def test_pipeline_raises_clear_error_for_missing_input(tmp_path):
    missing = tmp_path / "does_not_exist.csv"
    with pytest.raises(FileNotFoundError, match="Input file not found"):
        run(src=missing, dst=tmp_path / "out.xlsx")


def test_pipeline_handles_all_rows_dropped_without_crashing(tmp_path):
    dst = tmp_path / "out.xlsx"

    counts = run(src=ALL_DROPPED_FIXTURE, dst=dst)

    assert sum(counts.values()) == 0
    assert dst.exists()

    wb = load_workbook(dst)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert ws.max_row == 2  # title row + header row only, no data rows


def test_pipeline_creates_missing_output_parent_directories(tmp_path):
    dst = tmp_path / "reports" / "2026" / "out.xlsx"

    run(src=FIXTURE, dst=dst)

    assert dst.exists()


def test_pipeline_forwards_on_progress_to_workbook_build(tmp_path):
    dst = tmp_path / "out.xlsx"
    calls = []

    counts = run(src=FIXTURE, dst=dst, on_progress=lambda done, total: calls.append((done, total)))

    assert calls, "on_progress should fire at least once (the initial (0, total) call)"
    assert calls[0][0] == 0
    assert calls[-1][0] == calls[-1][1] == sum(counts.values())


def test_pipeline_writes_merged_title_row_above_header(tmp_path):
    dst = tmp_path / "out.xlsx"
    run(src=FIXTURE, dst=dst)

    wb = load_workbook(dst)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert ws.cell(row=1, column=1).value in {
            "POs Received Within Reporting Month",
            "POs Received Outside Reporting Month",
        }
        last_col_letter = get_column_letter(ws.max_column)
        assert f"A1:{last_col_letter}1" in {str(r) for r in ws.merged_cells.ranges}
