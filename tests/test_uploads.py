from pathlib import Path

import pytest
from openpyxl import load_workbook

from src import config
from src.uploads import (
    InvalidUpload,
    load_raw_preview,
    process_upload,
    validate_columns,
    validate_extension,
)

SAMPLE = Path(__file__).parent.parent / "data" / "sample" / "sample_raw_data.csv"
WRONG_COLUMNS = Path(__file__).parent / "fixtures" / "wrong_columns.csv"
HEADER_ONLY = Path(__file__).parent / "fixtures" / "header_only.csv"


def test_validate_extension_rejects_non_csv_filename():
    with pytest.raises(InvalidUpload):
        validate_extension("data.txt")


def test_validate_extension_accepts_csv_case_insensitively():
    validate_extension("data.csv")
    validate_extension("DATA.CSV")


def test_validate_columns_rejects_missing_business_columns():
    header = ["vend", "reqdate", "po"]
    with pytest.raises(InvalidUpload, match="tlc"):
        validate_columns(header)


def test_validate_columns_accepts_full_header():
    header = [*config.COLUMNS_TO_KEEP, "some_extra_column"]
    validate_columns(header)


def test_load_raw_preview_returns_header_and_limited_rows():
    csv_bytes = SAMPLE.read_bytes()

    header, rows, dropped_row_count = load_raw_preview(csv_bytes, limit=2)

    assert "vend" in header
    assert "reqhdr_id" in header  # raw header, not the trimmed business set
    assert len(rows) <= 2
    assert dropped_row_count == 0


def test_load_raw_preview_raises_invalid_upload_for_missing_columns():
    csv_bytes = WRONG_COLUMNS.read_bytes()

    with pytest.raises(InvalidUpload, match="tlc"):
        load_raw_preview(csv_bytes)


def test_load_raw_preview_raises_value_error_on_empty_bytes():
    with pytest.raises(ValueError, match="empty"):
        load_raw_preview(b"")


def test_process_upload_returns_xlsx_bytes_and_counts():
    csv_bytes = SAMPLE.read_bytes()

    result = process_upload(csv_bytes)

    assert result["xlsx_bytes"][:2] == b"PK"
    assert sum(result["counts"].values()) > 0
    assert set(result["preview_sheets"].keys()) == {
        config.WITHIN_SHEET,
        config.OUTSIDE_SHEET,
    }
    assert result["dropped_row_count"] == 0
    assert result["skipped_row_count"] == 0


def test_process_upload_xlsx_bytes_load_with_openpyxl(tmp_path):
    csv_bytes = SAMPLE.read_bytes()

    result = process_upload(csv_bytes)

    out = tmp_path / "out.xlsx"
    out.write_bytes(result["xlsx_bytes"])
    wb = load_workbook(out)
    assert set(wb.sheetnames) <= {config.WITHIN_SHEET, config.OUTSIDE_SHEET}


def test_process_upload_cleans_up_temp_directory():
    import tempfile

    captured_dirs = []
    real_tempdir = tempfile.TemporaryDirectory

    class TrackingTemporaryDirectory(real_tempdir):
        def __enter__(self):
            name = super().__enter__()
            captured_dirs.append(name)
            return name

    tempfile.TemporaryDirectory = TrackingTemporaryDirectory
    try:
        process_upload(SAMPLE.read_bytes())
    finally:
        tempfile.TemporaryDirectory = real_tempdir

    assert captured_dirs, "process_upload should use tempfile.TemporaryDirectory"
    assert not Path(captured_dirs[0]).exists()


def test_process_upload_calls_on_stage_with_expected_stage_sequence():
    csv_bytes = SAMPLE.read_bytes()
    stages = []

    process_upload(csv_bytes, on_stage=stages.append)

    assert stages == ["reading", "preparing", "building_report"]


def test_process_upload_calls_on_progress_during_building_report():
    csv_bytes = SAMPLE.read_bytes()
    calls = []

    result = process_upload(csv_bytes, on_progress=lambda done, total: calls.append((done, total)))

    assert calls, "on_progress should fire at least once (the initial (0, total) call)"
    assert calls[0][0] == 0
    assert calls[-1][0] == calls[-1][1] == sum(result["counts"].values())


def test_process_upload_without_on_stage_behaves_unchanged():
    csv_bytes = SAMPLE.read_bytes()

    result = process_upload(csv_bytes)  # on_stage omitted, as CLI/older callers do

    assert result["xlsx_bytes"][:2] == b"PK"


def test_process_upload_skips_rows_with_malformed_receive_date_and_reports_count():
    bad_csv = (
        ",".join(config.COLUMNS_TO_KEEP) + "\r\n"
        + "Acme,1/1/2026,PO1,1/5/2026,INV1,1/10/2026,NOT-A-DATE,100.00,IT,TLC1,Jane,Doe\r\n"
    ).encode(config.ENCODING)

    result = process_upload(bad_csv)

    assert result["xlsx_bytes"][:2] == b"PK"
    assert result["skipped_row_count"] == 1
    assert sum(result["counts"].values()) == 0
