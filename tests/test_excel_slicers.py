import xml.dom.minidom as minidom
import zipfile

from openpyxl import load_workbook

from src.config import COLUMNS_TO_KEEP, OUTSIDE_SHEET, WITHIN_SHEET
from src.excel.excel_slicers import SheetSlicerTarget, add_reporting_month_slicers
from src.excel.workbook_builder import build_workbook

HEADER = ["reporting_month"] + COLUMNS_TO_KEEP


def _row(reporting_month="January 2026", tlc="1/15/2026", receive_date="1/20/2026 00:00:00",
         vend="Acme Co", po="PO123", dept="IT"):
    return [reporting_month, vend, "1/1/2026", po, "1/10/2026", "INV1", tlc,
            receive_date, 100.00, dept, "TLC1", "Jane", "Doe"]


def _saved_contents(tmp_path, sheets):
    """A real, openpyxl-saved zip part dict - the same shape save_workbook
    passes into add_reporting_month_slicers, but without its own
    ignoredErrors/slicer patching applied yet."""
    wb, _, slicer_info = build_workbook(HEADER, sheets)
    base = tmp_path / "base.xlsx"
    wb.save(base)
    with zipfile.ZipFile(base) as zf:
        contents = {name: zf.read(name) for name in zf.namelist()}
    return contents, slicer_info


def _target_for(contents, slicer_info, sheet_name):
    table_id, reporting_month_column, num_columns = slicer_info[sheet_name]
    sheet_path = f"xl/worksheets/sheet{table_id}.xml"
    assert sheet_path in contents, f"expected {sheet_path} in saved zip parts"
    return SheetSlicerTarget(
        sheet_path=sheet_path,
        table_id=table_id,
        slicer_index=1,
        reporting_month_column=reporting_month_column,
        num_columns=num_columns,
    )


def test_add_reporting_month_slicers_adds_well_formed_xml_parts(tmp_path):
    contents, slicer_info = _saved_contents(tmp_path, {WITHIN_SHEET: [_row()]})
    target = _target_for(contents, slicer_info, WITHIN_SHEET)

    patched = add_reporting_month_slicers(contents, [target])

    for part in ("xl/slicerCaches/slicerCache1.xml", "xl/slicers/slicer1.xml", "xl/drawings/drawing1.xml"):
        assert part in patched
        minidom.parseString(patched[part])  # raises ExpatError if malformed


def test_add_reporting_month_slicers_references_correct_table_and_column(tmp_path):
    contents, slicer_info = _saved_contents(tmp_path, {WITHIN_SHEET: [_row()], OUTSIDE_SHEET: [_row()]})
    target = _target_for(contents, slicer_info, OUTSIDE_SHEET)

    patched = add_reporting_month_slicers(contents, [target])

    cache_xml = patched["xl/slicerCaches/slicerCache1.xml"].decode("utf-8")
    assert f'tableId="{target.table_id}"' in cache_xml
    assert f'column="{target.reporting_month_column}"' in cache_xml


def test_add_reporting_month_slicers_registers_content_type_overrides(tmp_path):
    contents, slicer_info = _saved_contents(tmp_path, {WITHIN_SHEET: [_row()]})
    target = _target_for(contents, slicer_info, WITHIN_SHEET)

    patched = add_reporting_month_slicers(contents, [target])

    content_types = patched["[Content_Types].xml"].decode("utf-8")
    assert "slicerCache1.xml" in content_types
    assert "slicer1.xml" in content_types
    assert "drawing1.xml" in content_types

    workbook_rels = patched["xl/_rels/workbook.xml.rels"].decode("utf-8")
    assert "slicerCaches/slicerCache1.xml" in workbook_rels


def test_add_reporting_month_slicers_wires_drawing_and_slicer_into_the_worksheet(tmp_path):
    contents, slicer_info = _saved_contents(tmp_path, {WITHIN_SHEET: [_row()]})
    target = _target_for(contents, slicer_info, WITHIN_SHEET)

    patched = add_reporting_month_slicers(contents, [target])

    sheet_xml = patched[target.sheet_path].decode("utf-8")
    assert "<drawing " in sheet_xml
    assert "<x14:slicerList>" in sheet_xml

    folder, filename = target.sheet_path.rsplit("/", 1)
    sheet_rels_path = f"{folder}/_rels/{filename}.rels"
    assert sheet_rels_path in patched
    sheet_rels = patched[sheet_rels_path].decode("utf-8")
    assert "../drawings/drawing1.xml" in sheet_rels
    assert "../slicers/slicer1.xml" in sheet_rels


def test_add_reporting_month_slicers_does_not_corrupt_the_zip(tmp_path):
    contents, slicer_info = _saved_contents(tmp_path, {WITHIN_SHEET: [_row()]})
    target = _target_for(contents, slicer_info, WITHIN_SHEET)

    patched = add_reporting_month_slicers(contents, [target])

    dst = tmp_path / "patched.xlsx"
    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in patched.items():
            zout.writestr(name, data)

    with zipfile.ZipFile(dst) as zin:
        assert zin.testzip() is None

    wb = load_workbook(dst)  # must not raise
    assert wb.sheetnames == [WITHIN_SHEET]
