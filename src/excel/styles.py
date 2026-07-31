"""openpyxl style objects used to format the report workbook.

Split from config.py because these are Style/Font/Fill objects (openpyxl
types), not plain business config.
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Subtle, mutually distinguishable row-shading colors, cycled per
# reporting_month section.
MONTH_FILL_COLORS = [
    "DDEBF7",  # light blue
    "E2EFDA",  # light green
    "FFF2CC",  # light yellow
    "FCE4D6",  # light orange
    "EAD1DC",  # light pink
    "E4DFEC",  # light purple
    "D9E1F2",  # light indigo
    "F2F2F2",  # light grey
]

ACCENT_COLOR = "1F4E78"
HIGHLIGHT_FONT = Font(bold=True, color=ACCENT_COLOR)
HIGHLIGHT_DATE_FORMAT = "mm/dd/yyyy"
HIGHLIGHT_ALIGNMENT = Alignment(horizontal="center")
HEADER_HIGHLIGHT_FILL = PatternFill("solid", fgColor=ACCENT_COLOR)
HEADER_HIGHLIGHT_FONT = Font(bold=True, color="FFFFFF")

# Bold, highly saturated border color chosen to read clearly against every
# color in MONTH_FILL_COLORS (all light pastels) as well as the navy header.
HIGHLIGHT_BORDER_COLOR = "C00000"
# Fill for tlc/receive_date data cells: a fixed, distinct tint (not one of
# MONTH_FILL_COLORS) in the same red family as the border, so these two
# columns read as one consistent block regardless of the row's month color.
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FCE4E4")
HIGHLIGHT_BORDER = Border(
    left=Side(style="thick", color=HIGHLIGHT_BORDER_COLOR),
    right=Side(style="thick", color=HIGHLIGHT_BORDER_COLOR),
    top=Side(style="thick", color=HIGHLIGHT_BORDER_COLOR),
    bottom=Side(style="thick", color=HIGHLIGHT_BORDER_COLOR),
)

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FONT = Font(bold=True, color="000000")
HEADER_BORDER = Border(bottom=Side(style="thin", color="808080"))
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# Sheet title (merged across the table's columns, row above the header) -
# deliberately distinct from HEADER_FONT/HEADER_FILL so it reads as a title
# rather than an extra header row: larger, no fill.
TITLE_FONT = Font(bold=True, size=14, color=ACCENT_COLOR)
TITLE_ALIGNMENT = Alignment(horizontal="center", vertical="center")

MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 40
