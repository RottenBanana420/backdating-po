"""Builds the styled openpyxl Workbook: one sheet per reporting-period
bucket, sorted and shaded by reporting_month, with tlc/receive_date
highlighted and a real Excel Table (backing the Slicer added later by
workbook_writer.save_workbook).
"""
import logging

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

from ..config import HIGHLIGHT_COLUMNS, OUTSIDE_SHEET, OUTSIDE_TITLE, TEXT_COLUMNS, WITHIN_SHEET, WITHIN_TITLE
from ..logging_config import log_stage
from ..transform import parse_reporting_month, to_date_only
from .styles import (
    HEADER_ALIGNMENT,
    HEADER_BORDER,
    HEADER_FILL,
    HEADER_FONT,
    HEADER_HIGHLIGHT_FILL,
    HEADER_HIGHLIGHT_FONT,
    HIGHLIGHT_ALIGNMENT,
    HIGHLIGHT_BORDER,
    HIGHLIGHT_DATE_FORMAT,
    HIGHLIGHT_FILL,
    HIGHLIGHT_FONT,
    MAX_COLUMN_WIDTH,
    MIN_COLUMN_WIDTH,
    MONTH_FILL_COLORS,
    TITLE_ALIGNMENT,
    TITLE_FONT,
)

logger = logging.getLogger(__name__)

# Row layout: a merged title row, then the column header row, then data -
# named here so every row-position reference below (the Table's ref, the
# ignoredErrors sqref workbook_writer.py patches in) stays in lockstep if
# this layout ever changes again.
TITLE_ROW = 1
HEADER_ROW = 2
DATA_START_ROW = 3

SHEET_TITLES = {WITHIN_SHEET: WITHIN_TITLE, OUTSIDE_SHEET: OUTSIDE_TITLE}

# Leading characters that Excel (and openpyxl) treat as formula triggers -
# OWASP's CSV Injection guidance: https://owasp.org/www-community/attacks/CSV_Injection
FORMULA_INJECTION_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")

# openpyxl's write-only mode has no built-in progress hooks (there's nothing
# to poll - cells stream straight to disk), and the per-cell styling loop
# below is this pipeline's dominant cost, so without an explicit callback a
# caller sees one "started" notification and then silence until the whole
# workbook is done - at 100k rows that's roughly a minute of an apparently
# frozen UI. Reporting every N rows instead of every row keeps the callback
# overhead negligible while still giving a UI something to show every couple
# of seconds.
DEFAULT_PROGRESS_INTERVAL_ROWS = 2000

TEXT_ALIGNMENT = Alignment(horizontal="left")


def _defuse_formula_injection(cell):
    """Defuses formula-injection payloads (e.g. "=HYPERLINK(...)") in a
    free-text cell pulled straight from the raw CSV, without altering the
    value the report displays.

    openpyxl only auto-marks a cell as a live formula (data_type='f') for a
    leading "=", but Excel's own cell editor re-parses +/-/@-prefixed text
    into a formula the moment a user opens and re-commits the cell (F2,
    Enter) - so all four OWASP-listed triggers need defusing, not just "=".
    Setting quotePrefix (the same style flag Excel itself sets when a user
    types a leading apostrophe) forces the cell to render and re-edit as
    plain text while leaving `cell.value` exactly as sourced - no visible
    apostrophe added to a legitimate value that happens to start with one
    of these characters (e.g. a vendor name like "-A1 Logistics").
    """
    if isinstance(cell.value, str) and cell.value.startswith(FORMULA_INJECTION_TRIGGERS):
        cell.data_type = "s"
        cell.quotePrefix = True


# log_stage is built on contextlib.contextmanager, which Python allows to
# double as a decorator - used that way here instead of a `with` block so
# this ~100-line function body doesn't need an extra indent level.
@log_stage(logger, "build styled workbook (shading, headers, tables)")
def build_workbook(header, sheets, on_progress=None, progress_interval_rows=DEFAULT_PROGRESS_INTERVAL_ROWS):
    """`on_progress`, if given, is called with (rows_done, rows_total) - once
    up front with rows_done=0, every `progress_interval_rows` data rows
    written, and once more at the end of each sheet so the final call always
    lands exactly on rows_total rather than drifting past it. rows_total is
    fixed for the whole build, computed once from `sheets`, so a caller can
    turn this straight into a 0-1 fraction without tracking sheet counts
    itself.
    """
    # write_only streams each row straight to disk instead of materializing
    # every cell as a full-blown openpyxl Cell object in an in-memory grid -
    # at 100k+ rows the per-cell style assignments were ~76% of this
    # pipeline's runtime. The tradeoff is that a write-only worksheet is
    # append-only: cells can't be styled or read back after the fact, so
    # every style below is built into each WriteOnlyCell before it's
    # appended, and structural, sheet-level properties (column widths,
    # freeze_panes) must be set before the first row goes out.
    wb = Workbook(write_only=True)

    rows_total = sum(len(rows) for rows in sheets.values())
    rows_done = 0

    def _report_progress():
        if on_progress is not None:
            on_progress(rows_done, rows_total)

    _report_progress()

    reporting_month_idx = header.index("reporting_month")
    receive_date_idx = header.index("receive_date")
    highlight_indices = {header.index(c) for c in HIGHLIGHT_COLUMNS}
    # reporting_month / po / invoiceid must stay Text-formatted and
    # left-aligned so Excel doesn't auto-detect them and re-flag them.
    text_format_indices = {reporting_month_idx} | {header.index(c) for c in TEXT_COLUMNS}

    # sheet_name -> (text_column_letters, data_start_row, last_row), needed
    # to re-patch the "Number Stored as Text" ignoredErrors block after save
    # (openpyxl does not serialize it).
    text_column_letters_by_sheet = {}
    # sheet_name -> SheetSlicerTarget-ish info, needed to patch in a real
    # reporting_month Slicer after save (openpyxl can't create one itself -
    # see excel_slicers.py).
    slicer_info_by_sheet = {}

    for table_id, (sheet_name, rows) in enumerate(sheets.items(), start=1):
        logger.debug("Building sheet %r: %d row(s)", sheet_name, len(rows))
        ws = wb.create_sheet(sheet_name)

        # Primary sort by reporting_month, then by receive_date ascending
        # (oldest to newest) within each reporting_month group.
        data_rows = sorted(
            rows,
            key=lambda r: (parse_reporting_month(r[reporting_month_idx]), to_date_only(r[receive_date_idx])),
        )
        data_rows = [
            tuple(
                to_date_only(value) if col_num in highlight_indices else value
                for col_num, value in enumerate(row)
            )
            for row in data_rows
        ]
        last_row = len(data_rows) + HEADER_ROW

        # Column widths (auto-sized to each column's longest value) precede
        # the cell data in the worksheet XML, so write-only mode requires
        # setting them before any row is appended. A single pass over
        # data_rows tracking a running max per column avoids one full
        # O(rows) rescan per column.
        longest_by_col = [len(str(col_name)) for col_name in header]
        for row in data_rows:
            for col_idx, value in enumerate(row):
                if value is not None:
                    value_len = len(str(value))
                    if value_len > longest_by_col[col_idx]:
                        longest_by_col[col_idx] = value_len
        for col_idx, longest in enumerate(longest_by_col, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(
                MIN_COLUMN_WIDTH, min(longest + 2, MAX_COLUMN_WIDTH)
            )

        # Title row: merged across every table column and centered, styled
        # distinctly from the header row below it so it clearly reads as a
        # title rather than a second header. write_only worksheets have no
        # add_table-style merge_cells() (it dereferences ws._cells, which a
        # streaming worksheet never populates), so the merge range is added
        # directly to ws.merged_cells - the writer only ever reads that
        # collection back out to emit <mergeCells>, nothing else touches it.
        title_cell = WriteOnlyCell(ws, value=SHEET_TITLES.get(sheet_name, sheet_name))
        title_cell.font = TITLE_FONT
        title_cell.alignment = TITLE_ALIGNMENT
        ws.append([title_cell, *([None] * (len(header) - 1))])
        ws.merged_cells.add(f"A{TITLE_ROW}:{get_column_letter(len(header))}{TITLE_ROW}")

        # Header row: bold/centered/grey, with tlc/receive_date called out
        # in the accent color.
        header_cells = []
        for col_idx, col_name in enumerate(header, start=1):
            cell = WriteOnlyCell(ws, value=col_name)
            if col_idx - 1 in highlight_indices:
                cell.fill = HEADER_HIGHLIGHT_FILL
                cell.font = HEADER_HIGHLIGHT_FONT
                cell.border = HIGHLIGHT_BORDER
            else:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = HEADER_BORDER
            cell.alignment = HEADER_ALIGNMENT
            header_cells.append(cell)
        ws.append(header_cells)

        # Data rows: shaded by reporting_month, alternating colors each time
        # the reporting_month changes; tlc/receive_date are then styled on
        # top of the shading so they stand out consistently in every row.
        color_idx = -1
        prev_month = None
        for row in data_rows:
            month = row[reporting_month_idx]
            if month != prev_month:
                color_idx = (color_idx + 1) % len(MONTH_FILL_COLORS)
                prev_month = month
            fill = PatternFill("solid", fgColor=MONTH_FILL_COLORS[color_idx])

            row_cells = []
            for col_num, value in enumerate(row):
                cell = WriteOnlyCell(ws, value=value)
                if col_num in highlight_indices:
                    cell.fill = HIGHLIGHT_FILL
                    cell.font = HIGHLIGHT_FONT
                    cell.border = HIGHLIGHT_BORDER
                    cell.number_format = HIGHLIGHT_DATE_FORMAT
                    cell.alignment = HIGHLIGHT_ALIGNMENT
                else:
                    cell.fill = fill
                    if col_num in text_format_indices:
                        cell.number_format = "@"
                        cell.alignment = TEXT_ALIGNMENT
                _defuse_formula_injection(cell)
                row_cells.append(cell)
            ws.append(row_cells)

            rows_done += 1
            if rows_done % progress_interval_rows == 0:
                _report_progress()

        # Flush unconditionally at sheet end: a sheet whose row count isn't a
        # multiple of progress_interval_rows would otherwise never report its
        # last few rows, and the very last flush across all sheets is what
        # brings rows_done to exactly rows_total.
        _report_progress()

        text_column_letters = [get_column_letter(header.index(c) + 1) for c in TEXT_COLUMNS]
        text_column_letters_by_sheet[ws.title] = (text_column_letters, DATA_START_ROW, last_row)

        # Real Excel Table backing the reporting_month Slicer added below.
        # No built-in style/banding, so it doesn't fight with the month
        # shading and header styling already applied above. Starts at the
        # header row, one below the title row above it.
        table_ref = f"A{HEADER_ROW}:{get_column_letter(len(header))}{last_row}"
        table = Table(
            id=table_id,
            displayName=f"Table{table_id}",
            ref=table_ref,
            tableColumns=[TableColumn(id=i, name=str(name)) for i, name in enumerate(header, start=1)],
        )
        table.tableStyleInfo = TableStyleInfo(
            showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False
        )
        ws.add_table(table)
        slicer_info_by_sheet[ws.title] = (table_id, reporting_month_idx + 1, len(header))

    return wb, text_column_letters_by_sheet, slicer_info_by_sheet
